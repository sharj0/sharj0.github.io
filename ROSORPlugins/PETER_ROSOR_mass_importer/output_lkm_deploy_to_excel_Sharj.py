from openpyxl import Workbook
import os


#Creates a simple excel file with the summed travel distance in km
def create_excel_file(folder_path, total_lkm):
    #Initializes workbook with one sheet named total_lkm
    workbook_output_lkm = Workbook()
    workbook_output_lkm["Sheet"].title = "total_lkm"
    workspace = workbook_output_lkm["total_lkm"]

    # Indexes for cells start at 1 (need this as a note)
    # First column is the lkm data
    workspace.cell(2, 1).value = "Total LKM From csvs"
    workspace.cell(3, 1).value = f'{total_lkm:.3f}'

    #second column is the units (just as a quadruple reminder that it's in km)
    workspace.cell(2, 2).value = "Units"
    workspace.cell(3, 2).value = "km"

    #Uses the imput folder as tbe naming basis for the excel file  and inserts _output_lkm at the end to denote the value is for every csv in that folder
    save_file_name = os.path.basename(folder_path) + "_output_lkm.xlsx"
    directory_name = os.path.dirname(folder_path)
    output_file_save_path = os.path.join(directory_name, save_file_name)

    #Saves the file to the selected input folder (this can be changed)
    workbook_output_lkm.save(output_file_save_path)

    return output_file_save_path

#Made opening an Excel file a function in case it comes up again (is not necessary as its a one liner anyways)
def open_excel_file(excel_file_full_path):
    #os.startfile() will use the default program to run the excel file (makes a bit it more cross platform)
    os.startfile(excel_file_full_path)
